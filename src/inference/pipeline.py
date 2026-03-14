"""
pipeline.py — Local inference pipeline.

Full flow: Load CSV → Drift check → Score → Write Excel

Usage:
    python src/inference/pipeline.py --incoming data/incoming/new_patients.csv

Reads:  incoming CSV
        data/processed/reference.csv (for drift check)
        models/champion/ (exported model)
Writes: output/scored_YYYYMMDD_HHMMSS.xlsx
"""

import sys
import argparse
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

import pandas as pd
import numpy as np
import joblib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from src.config import (
    ALL_FEATURES, TARGET, MODEL_DIR, PROJECT_ROOT
)
from src.monitoring.drift_check import load_reference, check_drift, print_result


OUTPUT_DIR = PROJECT_ROOT / "output"
MODEL_PKL = MODEL_DIR / "champion" / "model.pkl"


def load_model():
    """Load the exported champion model."""
    if not MODEL_PKL.exists():
        raise FileNotFoundError(f"Model not found at {MODEL_PKL}. Run export_model.py first.")
    model = joblib.load(MODEL_PKL)
    print(f"Model loaded from: {MODEL_PKL}")
    return model


def score_batch(model, df):
    """Score a batch of records."""
    X = df[ALL_FEATURES]
    predictions = model.predict(X)
    probabilities = model.predict_proba(X)[:, 1]

    df = df.copy()
    df["prediction"] = predictions
    df["prediction_label"] = ["Stroke" if p == 1 else "No Stroke" for p in predictions]
    df["stroke_probability"] = np.round(probabilities, 4)
    df["confidence"] = np.round(np.maximum(probabilities, 1 - probabilities), 4)

    stroke_count = int(predictions.sum())
    print(f"\nScoring complete:")
    print(f"  Total records:  {len(df)}")
    print(f"  Stroke:         {stroke_count}")
    print(f"  No Stroke:      {len(df) - stroke_count}")
    print(f"  Avg confidence: {df['confidence'].mean():.4f}")

    return df


def write_excel(df, drift_result):
    """Write scored results to a formatted Excel file."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = OUTPUT_DIR / f"scored_{timestamp}.xlsx"

    wb = Workbook()

    # ── Sheet 1: Summary ──
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # Styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
    stroke_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    no_stroke_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Summary content
    summary_data = [
        ("Stroke Prediction — Batch Scoring Report", ""),
        ("", ""),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Total Records", len(df)),
        ("Stroke Predicted", int(df["prediction"].sum())),
        ("No Stroke Predicted", int((df["prediction"] == 0).sum())),
        ("Average Confidence", f"{df['confidence'].mean():.1%}"),
        ("", ""),
        ("Drift Check", ""),
        ("Features Checked", drift_result["n_total"]),
        ("Features Drifted", drift_result["n_drifted"]),
        ("Drift Share", f"{drift_result['drift_share']:.1%}"),
        ("Drift Threshold", f"{drift_result['threshold']:.1%}"),
        ("Decision", "PASS" if not drift_result["drift_detected"] else "FAIL"),
    ]

    for row_idx, (key, value) in enumerate(summary_data, 1):
        ws_summary.cell(row=row_idx, column=1, value=key).font = Font(bold=True)
        ws_summary.cell(row=row_idx, column=2, value=value)

    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 30

    # Title styling
    ws_summary.cell(row=1, column=1).font = Font(bold=True, size=14, color="2B579A")

    # ── Sheet 2: Scored Results ──
    ws_results = wb.create_sheet("Scored Results")

    # Headers
    columns = ALL_FEATURES + ["prediction", "prediction_label", "stroke_probability", "confidence"]
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws_results.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Data rows
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws_results.cell(row=row_idx, column=col_idx, value=row[col_name])
            cell.border = thin_border

            # Color code prediction
            if col_name == "prediction_label":
                if row[col_name] == "Stroke":
                    cell.fill = stroke_fill
                else:
                    cell.fill = no_stroke_fill

    # Auto-width columns
    for col_idx, col_name in enumerate(columns, 1):
        ws_results.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "A"].width = max(len(col_name) + 4, 12)

    # Save
    wb.save(str(output_path))
    print(f"\nExcel saved: {output_path}")

    return output_path


def run_pipeline(incoming_path, skip_drift=False):
    """
    Run the full inference pipeline.

    Steps:
        1. Load incoming data
        2. Run drift check
        3. If drift OK (or skip_drift) → score batch
        4. Write Excel report
    """
    print("=" * 60)
    print("STROKE PREDICTION — INFERENCE PIPELINE")
    print("=" * 60)

    # Step 1: Load incoming data
    incoming_df = pd.read_csv(incoming_path)
    print(f"\nStep 1: Loaded {len(incoming_df)} records from {incoming_path}")

    # Step 2: Drift check
    print(f"\nStep 2: Running drift check...")
    reference_df = load_reference()
    drift_result = check_drift(reference_df, incoming_df)
    print_result(drift_result)

    if drift_result["drift_detected"] and not skip_drift:
        print("\n⛔ PIPELINE HALTED — Drift exceeds threshold.")
        print("   Use --skip-drift to force scoring anyway.")
        return None

    # Step 3: Score
    print(f"\nStep 3: Scoring batch...")
    model = load_model()
    scored_df = score_batch(model, incoming_df)

    # Step 4: Write Excel
    print(f"\nStep 4: Writing Excel report...")
    output_path = write_excel(scored_df, drift_result)

    print(f"\n{'=' * 60}")
    print("PIPELINE COMPLETE")
    print(f"  Records scored: {len(scored_df)}")
    print(f"  Output:         {output_path}")
    print(f"{'=' * 60}")

    return output_path


def main():
    parser = argparse.ArgumentParser(description="Run inference pipeline")
    parser.add_argument("--incoming", required=True, help="Path to incoming CSV")
    parser.add_argument("--skip-drift", action="store_true", help="Skip drift check and score anyway")
    args = parser.parse_args()

    run_pipeline(args.incoming, skip_drift=args.skip_drift)


if __name__ == "__main__":
    main()